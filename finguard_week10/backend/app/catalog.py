"""
Catalog metadata loader.

Reads FinGuard_Data_Catalog.xlsx (98 columns) and enriches each row with
classification owner/steward, retention period, and regulatory reference
inferred from the Week 7 classification.
"""

from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List
import openpyxl


# ---- Static enrichment lookups (sourced from the handover doc & Week 7 sheet) ----

# Owner is the CDO for everything per the handover; stewards differ.
DEFAULT_OWNER = "CDO — Maria Chen"

STEWARD_BY_TABLE: Dict[str, str] = {
    "aml_transactions":          "AML Officer — Sarah Johnson",
    "bank_marketing_customers":  "DPO — Klaus Weber",
    "sec_edgar_filings":         "AML Officer — Sarah Johnson",
    "paysim_fraud_transactions": "Data Engineer — James Park",
    "ecb_statistical_data":      "Data Engineer — Li Wei",
    "world_bank_indicators":     "Data Engineer — Li Wei",
    "fatf_country_risk":         "AML Officer — Sarah Johnson",
}

# Default retention in years per classification (mirrors Week 7 sheet).
RETENTION_BY_CLASSIFICATION: Dict[str, str] = {
    "Public":              "10 years (regulatory reference)",
    "Internal":            "7 years (AML record-keeping)",
    "Confidential":        "5 years (AML/GLBA)",
    "Restricted":          "3 years post relationship (GDPR minimization)",
    "Highly Restricted":   "3 years post relationship (GDPR minimization)",
}

# Map common PII fields to the regulation that most directly governs them.
REGULATION_HINTS: Dict[str, str] = {
    "customer_name":      "GDPR Art. 4(1); GLBA",
    "full_name":          "GDPR Art. 4(1); GLBA",
    "ceo_name":           "GDPR Art. 4(1)",
    "beneficiary_name":   "GDPR Art. 4(1)",
    "name_orig":          "GDPR Art. 4(5) (pseudonymization)",
    "name_dest":          "GDPR Art. 4(5) (pseudonymization)",
    "email":              "GDPR Art. 4(1); CAN-SPAM",
    "customer_email":     "GDPR Art. 4(1)",
    "phone":              "GDPR Art. 4(1); GLBA",
    "customer_phone":     "GDPR Art. 4(1); GLBA",
    "address":            "GDPR Art. 4(1)",
    "customer_address":   "GDPR Art. 4(1)",
    "headquarters_address": "GDPR Art. 4(1)",
    "date_of_birth":      "GDPR Art. 9 (special category-adjacent)",
    "customer_dob":       "GDPR Art. 9",
    "age":                "EU AI Act (bias-sensitive feature)",
    "ip_address":         "GDPR Recital 30",
    "device_id":          "GDPR Art. 4(1)",
    "originator_account": "PSD2; GDPR",
    "beneficiary_account":"PSD2; GDPR",
    "ein":                "GLBA",
    "zip_code":           "GDPR Art. 4(1)",
    "marital":            "GDPR Art. 9",
    "country":            "GDPR (cross-border transfer)",
    "originator_country": "GDPR (cross-border transfer)",
    "beneficiary_country":"GDPR (cross-border transfer)",
    "customer_id":        "GDPR Art. 4(5)",
    "risk_score":         "EU AI Act (high-risk system output)",
    "aml_flag":           "EU AMLD",
    "is_suspicious":      "EU AMLD",
    "is_fraud":           "Fraud detection — internal use",
}

# Access tier follows the classification:
ACCESS_BY_CLASSIFICATION: Dict[str, str] = {
    "Public":             "Public",
    "Internal":           "Internal",
    "Confidential":       "Restricted",
    "Restricted":         "Highly Restricted",
    "Highly Restricted":  "Highly Restricted",
}


def _dq_summary(row: Dict[str, Any]) -> Dict[str, bool]:
    return {
        "missing":     row.get("DQ: Missing") == "Y",
        "invalid":     row.get("DQ: Invalid") == "Y",
        "inconsistent":row.get("DQ: Inconsistent") == "Y",
        "duplicate":   row.get("DQ: Duplicate") == "Y",
        "out_of_range":row.get("DQ: Out of Range") == "Y",
        "format_error":row.get("DQ: Format Error") == "Y",
    }


def load_catalog(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read the master catalog spreadsheet and enrich it."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Data Catalog"]
    headers = [c.value for c in ws[1]]
    out: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        table = r["Table Name"]
        col = r["Column Name"]
        classification = r.get("Classification") or "Internal"
        masking = r.get("Masking Method") or "-"
        is_pii = (r.get("PII?") == "Y")

        item = {
            "table": table,
            "column": col,
            "data_type": r.get("Data Type"),
            "sample_value": r.get("Sample Value"),
            "is_pii": is_pii,
            "pii_type": r.get("PII Type") or "-",
            "classification": classification,
            "dq": _dq_summary(r),
            "dq_any_issue": any(_dq_summary(r).values()),
            "masking_method": masking,
            "masking_required": masking not in (None, "-", ""),
            "owner": DEFAULT_OWNER,
            "steward": STEWARD_BY_TABLE.get(table, "Data Engineer — James Park"),
            "retention": RETENTION_BY_CLASSIFICATION.get(classification, "5 years"),
            "access_tier": ACCESS_BY_CLASSIFICATION.get(classification, "Internal"),
            "regulation": REGULATION_HINTS.get(col, "—" if not is_pii else "GDPR Art. 4(1)"),
        }
        out.append(item)
    return out


def summarize(catalog: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Aggregate stats used by the dashboard."""
    by_class: Dict[str, int] = {}
    by_pii: Dict[str, int] = {"Direct": 0, "Quasi": 0, "Sensitive": 0, "None": 0}
    by_table: Dict[str, int] = {}
    masked_pii = 0
    dq_with_issue = 0

    for c in catalog:
        by_class[c["classification"]] = by_class.get(c["classification"], 0) + 1
        if c["is_pii"]:
            ptype = c["pii_type"] if c["pii_type"] in ("Direct", "Quasi", "Sensitive") else "Direct"
            by_pii[ptype] = by_pii.get(ptype, 0) + 1
            if c["masking_required"]:
                masked_pii += 1
        else:
            by_pii["None"] += 1
        by_table[c["table"]] = by_table.get(c["table"], 0) + 1
        if c["dq_any_issue"]:
            dq_with_issue += 1

    return {
        "total_columns": len(catalog),
        "total_tables": len(by_table),
        "by_classification": by_class,
        "by_pii_type": by_pii,
        "by_table": by_table,
        "pii_columns": sum(1 for c in catalog if c["is_pii"]),
        "pii_masked_columns": masked_pii,
        "dq_columns_with_issue": dq_with_issue,
    }
